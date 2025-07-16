"""
Script para automatizar a importação de PDFs de fichas financeiras.

Funcionalidades:
1. Lê todos os arquivos PDF de um diretório;
2. Extrai nome, CPF, data de admissão e a tabela de proventos/descontos;

Requisitos:
pip install pandas PyPDF2
"""

import os
import re
import pandas as pd
from PyPDF2 import PdfReader
from typing import Tuple, List, Optional
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pdfplumber

# Caminhos
PASTA_PDFS = "./pdfs"  # Coloque os PDFs nesta pasta

def extrair_texto_pdf(pdf_path: str) -> str:
    """
    Extrai todo o texto de um arquivo PDF.
    
    Args:
        pdf_path: Caminho para o arquivo PDF
        
    Returns:
        Texto extraído do PDF
    """
    try:
        reader = PdfReader(pdf_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n\n"
        return text
    except Exception as e:
        print(f"Erro ao ler PDF {pdf_path}: {e}")
        return ""

def extrair_informacoes_basicas(texto: str) -> Tuple[str, str, str]:
    """
    Extrai nome, CPF e data de admissão do texto do PDF.
    
    Args:
        texto: Texto extraído do PDF
        
    Returns:
        Tupla com (nome, cpf, admissao)
    """
    # Padrões para extração
    padrao_nome = r'MATRICULA:\s*\d+\s*-\s*(.+?)(?:\s+ADMISSÃO|$)'
    padrao_cpf = r'CPF:\s*([\d\.-]+)'
    padrao_admissao = r'ADMISSÃO:\s*(\d{2}/\d{2}/\d{4})'
    
    # Extrair informações
    nome_match = re.search(padrao_nome, texto, re.IGNORECASE)
    cpf_match = re.search(padrao_cpf, texto, re.IGNORECASE)
    admissao_match = re.search(padrao_admissao, texto, re.IGNORECASE)
    
    nome = nome_match.group(1).strip() if nome_match else "NOME NÃO ENCONTRADO"
    cpf = cpf_match.group(1).strip() if cpf_match else "CPF NÃO ENCONTRADO"
    admissao = admissao_match.group(1).strip() if admissao_match else "ADMISSÃO NÃO ENCONTRADA"
    
    return nome, cpf, admissao

def extrair_tabela_pdfplumber(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                # Procura por uma tabela que tenha o cabeçalho esperado
                if table and 'PROVENTOS/DESCONTO' in table[0][0]:
                    df = pd.DataFrame(table[1:], columns=table[0])
                    return df
    return None

def processar_pdf(pdf_path: str) -> dict:
    """
    Processa um arquivo PDF e extrai todas as informações necessárias.
    
    Args:
        pdf_path: Caminho para o arquivo PDF
        
    Returns:
        Dicionário com as informações extraídas
    """
    print(f"\nProcessando: {os.path.basename(pdf_path)}")
    
    # Extrair texto do PDF
    texto = extrair_texto_pdf(pdf_path)
    if not texto:
        return {"erro": "Não foi possível extrair texto do PDF"}
    
    # Extrair informações básicas
    nome, cpf, admissao = extrair_informacoes_basicas(texto)
    
    # Extrair tabela de proventos/descontos
    tabela = extrair_tabela_pdfplumber(pdf_path)
    
    # Detectar ano do PDF
    ano_match = re.search(r'COMPETÊNCIA\s*-\s*(\d{4})', texto, re.IGNORECASE)
    ano = ano_match.group(1) if ano_match else "AnoDesconhecido"

    resultado = {
        "arquivo": os.path.basename(pdf_path),
        "nome": nome,
        "cpf": cpf,
        "admissao": admissao,
        "ano": ano,
        "tabela": tabela,
        "texto_completo": texto[:500] + "..." if len(texto) > 500 else texto  # Primeiros 500 chars para debug
    }
    
    print(f"  Nome: {nome}")
    print(f"  CPF: {cpf}")
    print(f"  Admissão: {admissao}")
    print(f"  Ano: {ano}")
    print(f"  Tabela extraída: {'Sim' if tabela is not None else 'Não'}")
    
    return resultado

def processar_todos_pdfs() -> List[dict]:
    """
    Processa todos os arquivos PDF no diretório especificado.
    
    Returns:
        Lista de dicionários com os resultados de cada PDF
    """
    resultados = []
    
    # Verificar se a pasta existe
    if not os.path.exists(PASTA_PDFS):
        print(f"Pasta {PASTA_PDFS} não encontrada!")
        return resultados
    
    # Listar arquivos PDF
    arquivos_pdf = [f for f in os.listdir(PASTA_PDFS) if f.lower().endswith('.pdf')]
    
    if not arquivos_pdf:
        print(f"Nenhum arquivo PDF encontrado em {PASTA_PDFS}")
        return resultados
    
    print(f"Encontrados {len(arquivos_pdf)} arquivo(s) PDF:")
    for arquivo in arquivos_pdf:
        print(f"  - {arquivo}")
    
    # Processar cada PDF
    for arquivo in arquivos_pdf:
        pdf_path = os.path.join(PASTA_PDFS, arquivo)
        resultado = processar_pdf(pdf_path)
        resultados.append(resultado)
    
    return resultados

CAMINHO_PLANILHA = "./Cálculo.v15 - Poupança - Preenchido.xlsx"

import re

def split_nome_valores(linha_str, colunas):
    import re
    # Ignorar cabeçalho
    if linha_str.strip().lower().startswith(("janeiro", "proventos/desconto")):
        return None
    # Tentar split por múltiplos espaços
    partes = re.split(r'\s{2,}', linha_str)
    if len(partes) == len(colunas):
        return partes
    # Fallback: regex para valores monetários
    valores = re.findall(r'-?\d{1,3}(?:[\.\d]{0,3})*,\d{2}', linha_str)
    if valores:
        idx = linha_str.find(valores[0])
        nome = linha_str[:idx].strip()
        valores_str = linha_str[idx:].strip()
        valores = re.findall(r'-?\d{1,3}(?:[\.\d]{0,3})*,\d{2}', valores_str)
        if len(valores) == len(colunas) - 1:
            return [nome] + valores
        elif len(valores) < len(colunas) - 1:
            return [nome] + valores + [''] * (len(colunas) - 1 - len(valores))
    # Se não conseguir, retorna nome e colunas vazias
    return [linha_str.strip()] + [''] * (len(colunas) - 1)

def adicionar_tabela_excel(planilha_path, ano, tabela, nome, cpf, admissao):
    wb = openpyxl.load_workbook(planilha_path)
    # Remover aba do ano se já existir
    if ano in wb.sheetnames:
        std = wb[ano]
        wb.remove(std)
    # Criar nova aba do ano
    ws = wb.create_sheet(title=ano)
    # Escrever cabeçalho manualmente para garantir ordem correta
    colunas = ['PROVENTOS/DESCONTO', 'Janeiro', 'Fevereiro', 'Março', 'Abril',
               'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro',
               'Novembro', 'Dezembro', 'Total']
    ws.append(colunas)
    # Escrever dados corrigindo a separação do nome e valores
    for _, row in tabela.iterrows():
        linha_str = ' '.join([str(x) for x in row if pd.notnull(x)])
        resultado = split_nome_valores(linha_str, colunas)
        if resultado and len(resultado) == len(colunas): # 14 meses + 1 total
            ws.append(resultado)
    # (Opcional) Preencher dados em uma aba padrão, ex: 'Recalculo'
    if 'Recalculo' in wb.sheetnames:
        ws_recalculo = wb['Recalculo']
        ws_recalculo['C2'] = nome
        ws_recalculo['C3'] = cpf
        ws_recalculo['C4'] = admissao
    wb.save(planilha_path)
    print(f"Dados adicionados na aba '{ano}' da planilha.")

def main():
    """
    Função principal que executa o processamento.
    """
    print("=== AUTOMAÇÃO DE PROCESSAMENTO DE FICHAS FINANCEIRAS ===")
    print("Regras implementadas:")
    print("1. Lê todos os arquivos PDF de um diretório")
    print("2. Extrai nome, CPF, data de admissão e a tabela de proventos/descontos")
    print("3. Adiciona os dados extraídos à planilha Excel preenchida\n")
    
    # Processar todos os PDFs
    resultados = processar_todos_pdfs()
    
    # Resumo dos resultados
    print(f"\n=== RESUMO ===")
    print(f"Total de PDFs processados: {len(resultados)}")
    
    for resultado in resultados:
        if "erro" not in resultado:
            print(f"\nArquivo: {resultado['arquivo']}")
            print(f"  Nome: {resultado['nome']}")
            print(f"  CPF: {resultado['cpf']}")
            print(f"  Admissão: {resultado['admissao']}")
            print(f"  Ano: {resultado['ano']}")
            if resultado['tabela'] is not None:
                print(f"  Tabela: {len(resultado['tabela'])} linhas extraídas")
                adicionar_tabela_excel(
                    CAMINHO_PLANILHA,
                    resultado['ano'],
                    resultado['tabela'],
                    resultado['nome'],
                    resultado['cpf'],
                    resultado['admissao']
                )
            else:
                print("  Tabela: Não extraída")
        else:
            print(f"\nErro no arquivo {resultado.get('arquivo', 'desconhecido')}: {resultado['erro']}")
    
    print("\nProcessamento concluído!")

if __name__ == "__main__":
    main()
